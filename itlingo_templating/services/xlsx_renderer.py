"""Render an XLSX template against the canonical model with openpyxl + Jinja2.

The XLSX templates use the same Jinja2 syntax as the DOCX ones. openpyxl does
not understand Jinja, so this module interprets a restricted profile:

* Scalar/inline cells: any cell text with ``{{ ... }}`` (or inline
  ``{% if %}...{% endif %}``) is rendered in place.
* Repeated rows: a ``{% for VAR in EXPR %}`` row and a matching ``{% endfor %}``
  row (each alone on its own row) bracket one body row that is duplicated once
  per item, with VAR in scope. ``EXPR`` is a full Jinja expression, so standard
  filters/sorts work (``constraints | selectattr('type','equalto','Project')``).

Caveats (v1): no nested ``for``; avoid merged cells, charts and formulas inside
a loop region -- openpyxl does not adjust them when rows are inserted.
"""

import io
import re
from copy import copy

from .rendering import lenient_env

_FOR_RE = re.compile(r"^\s*\{%-?\s*for\s+(\w+)\s+in\s+(.+?)\s*-?%\}\s*$")
_ENDFOR_RE = re.compile(r"^\s*\{%-?\s*endfor\s*-?%\}\s*$")
_JINJA_RE = re.compile(r"\{\{|\{%")
# Integer-only on purpose: keeps counts (|length) numeric without mangling
# version-like strings ("1.0") or zero-padded codes into floats.
_INT_RE = re.compile(r"^-?\d+$")


def render_xlsx(template_bytes, context):
    """Render the XLSX *template_bytes* with *context*; return XLSX bytes."""
    from openpyxl import load_workbook

    env = lenient_env()
    wb = load_workbook(io.BytesIO(template_bytes))
    for ws in wb.worksheets:
        _expand_loops(ws, context, env)
        _render_scalar_cells(ws, context, env)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _first_cell_text(ws, row_idx):
    value = ws.cell(row=row_idx, column=1).value
    return value if isinstance(value, str) else ""


def _find_blocks(ws):
    """Return loop blocks as (for_row, body_row, endfor_row, var, expr).

    Blocks are returned bottom-to-top so callers can mutate rows without
    invalidating the indices of blocks they have not processed yet.
    """
    blocks = []
    row = 1
    max_row = ws.max_row
    while row <= max_row:
        match = _FOR_RE.match(_first_cell_text(ws, row))
        if not match:
            row += 1
            continue
        end = row + 1
        while end <= max_row and not _ENDFOR_RE.match(_first_cell_text(ws, end)):
            end += 1
        if end > max_row:
            break  # unmatched for: leave it untouched
        # Body is the single row between the markers (v1).
        blocks.append((row, row + 1, end, match.group(1), match.group(2)))
        row = end + 1
    blocks.reverse()
    return blocks


def _expand_loops(ws, context, env):
    for for_row, body_row, end_row, var, expr in _find_blocks(ws):
        try:
            items = list(env.compile_expression(expr)(**context) or [])
        except Exception:
            items = []

        template_cells = _snapshot_row(ws, body_row)
        # Remove the original body + the two marker rows; we re-insert below.
        ws.delete_rows(for_row, end_row - for_row + 1)

        for offset, item in enumerate(items):
            ws.insert_rows(for_row + offset)
            local = dict(context)
            local[var] = item
            _write_row(ws, for_row + offset, template_cells, local, env)


def _snapshot_row(ws, row_idx):
    """Capture (value, style, number_format) per cell of *row_idx*."""
    cells = []
    for cell in ws[row_idx]:
        cells.append((
            cell.column,
            cell.value,
            copy(cell._style),
            cell.number_format,
        ))
    return cells


def _write_row(ws, row_idx, template_cells, context, env):
    for column, value, style, number_format in template_cells:
        cell = ws.cell(row=row_idx, column=column)
        cell.value = _render_value(value, context, env)
        cell._style = copy(style)
        cell.number_format = number_format


def _render_scalar_cells(ws, context, env):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and _JINJA_RE.search(cell.value):
                cell.value = _render_value(cell.value, context, env)


def _render_value(value, context, env):
    if not isinstance(value, str) or not _JINJA_RE.search(value):
        return value
    try:
        rendered = env.from_string(value).render(**context)
    except Exception:
        return value
    return _coerce_number(rendered)


def _coerce_number(text):
    """Turn a purely integer rendered string into an int (e.g. counts)."""
    if not isinstance(text, str) or not _INT_RE.match(text):
        return text
    return int(text)

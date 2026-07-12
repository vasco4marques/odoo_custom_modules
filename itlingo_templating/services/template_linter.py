"""Static, advisory checks for Jinja expressions in document templates."""

import difflib
import io
import re
import zipfile

from .rendering import lenient_env
from .template_reference import BUILTIN_COMPATIBILITY_COLLECTIONS


GENERIC_VARIABLES = {"root", "elements", "by_type", "by_id", "schema_version"}
CATEGORY_ORDER = ("syntax", "filter", "test", "structure", "dsl")


def extract_template_sources(template_bytes, template_format):
    """Return Jinja-bearing sources and optional degraded top-level names."""
    template_format = (template_format or "").lower()
    if template_format == "docx":
        return _extract_docx(template_bytes)
    if template_format == "xlsx":
        return _extract_xlsx(template_bytes)
    return {
        "sources": [{
            "text": template_bytes.decode("utf-8"),
            "location": "line",
            "line_numbers": True,
        }],
        "top_level": set(),
        "degraded": False,
        "findings": [],
    }


def _extract_docx(template_bytes):
    from docxtpl import DocxTemplate

    convention_findings = _docx_structure_findings(template_bytes)
    template = DocxTemplate(io.BytesIO(template_bytes))
    try:
        template.init_docx()
        source = template.patch_xml(template.get_xml())
        return {
            "sources": [{"text": source, "location": None, "line_numbers": False}],
            "top_level": set(),
            "degraded": False,
            "findings": convention_findings,
        }
    except Exception:
        # docxtpl's own fallback can still discover names across headers,
        # footers, and other XML parts even when the main patched XML fails.
        try:
            names = template.get_undeclared_template_variables(
                jinja_env=lenient_env(),
            )
        except Exception:
            names = set()
        return {
            "sources": [],
            "top_level": set(names or ()),
            "degraded": True,
            "findings": convention_findings,
        }


def _extract_xlsx(template_bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(template_bytes), read_only=True, data_only=False)
    sources = []
    findings = []
    for sheet in workbook.worksheets:
        values = []
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ("{{" in cell.value or "{%" in cell.value):
                    values.append(cell.value)
                    cells.append(cell.coordinate)
                    findings.extend(_xlsx_structure_findings(
                        cell.value, sheet.title, cell.coordinate, cell.column,
                    ))
        if values:
            sources.append({
                "text": "\n".join(values),
                "location": "%s!%s" % (sheet.title, ",".join(cells[:5])),
                "line_numbers": False,
            })
    workbook.close()
    return {
        "sources": sources,
        "top_level": set(),
        "degraded": False,
        "findings": findings,
    }


def _docx_structure_findings(template_bytes):
    """Warn about the most reliably detectable docxtpl row-tag mismatch."""
    try:
        from lxml import etree

        with zipfile.ZipFile(io.BytesIO(template_bytes)) as archive:
            texts = []
            for name in archive.namelist():
                if name.startswith("word/") and name.endswith(".xml"):
                    root = etree.fromstring(archive.read(name))
                    texts.append("".join(root.itertext()))
    except Exception:
        return []
    text = "\n".join(texts)
    starts = len(re.findall(r"{%\s*tr\s+for\b", text))
    ends = len(re.findall(r"{%\s*tr\s+endfor\b", text))
    if starts == ends:
        return []
    return [_finding(
        "warning", "structure",
        "DOCX table-row loop markers are unbalanced (%s start, %s end)."
        % (starts, ends),
        "tr", location="DOCX table rows",
    )]


def _xlsx_structure_findings(value, sheet_name, coordinate, column):
    markers = re.findall(r"{%\s*(?:for\b[^%]*|endfor\s*)%}", value)
    if not markers:
        return []
    standalone = len(markers) == 1 and value.strip() == markers[0]
    if column == 1 and standalone:
        return []
    return [_finding(
        "info", "structure",
        "XLSX for/endfor markers should be alone in a column-A cell.",
        coordinate, location="%s!%s" % (sheet_name, coordinate),
    )]


def lint_template(template_bytes, template_format, reference_context, dsl_key=None):
    """Return Jinja and, when available, DSL-inventory findings."""
    extracted = extract_template_sources(template_bytes, template_format)
    dsl_available = bool(reference_context and reference_context.get("success"))
    inventory = _inventory(reference_context, dsl_key) if dsl_available else None
    findings = list(extracted.get("findings", ()))
    if inventory:
        for name in sorted(extracted["top_level"]):
            _check_top_level(name, inventory, findings, None)

    for source in extracted["sources"]:
        # Keep temporary placeholder filters/tests local to this extracted
        # source so findings and suggestions cannot leak between sheets/parts.
        env = lenient_env()
        try:
            ast = env.parse(source["text"])
        except Exception as err:
            from jinja2 import TemplateSyntaxError

            if not isinstance(err, TemplateSyntaxError):
                raise
            findings.append(_finding(
                "error", "syntax", "Invalid Jinja syntax: %s" % err.message,
                getattr(err, "name", None) or "jinja",
                location=_location(source, err),
            ))
            continue

        _check_filters_and_tests(ast, env, findings, source)
        if not inventory:
            continue
        from jinja2 import meta

        for name in sorted(meta.find_undeclared_variables(ast)):
            _check_top_level(name, inventory, findings, _location(source, None))
        _walk(ast, {}, inventory, findings, source)

    unique = []
    seen = set()
    for finding in findings:
        key = (
            finding["category"], finding["severity"], finding["symbol"],
            finding["message"], finding["location"],
        )
        if key not in seen:
            seen.add(key)
            unique.append(finding)
    rank = {category: index for index, category in enumerate(CATEGORY_ORDER)}
    unique.sort(key=lambda item: rank.get(item["category"], len(rank)))
    groups = [{
        "category": category,
        "label": category.title(),
        "findings": [item for item in unique if item["category"] == category],
    } for category in CATEGORY_ORDER if any(
        item["category"] == category for item in unique
    )]
    return {
        "skipped": False,
        "dsl_skipped": not dsl_available,
        "degraded": extracted["degraded"],
        "findings": unique,
        "finding_groups": groups,
    }


def _inventory(reference, dsl_key):
    type_items = {
        item.get("name"): item
        for item in reference.get("types", []) if item.get("name")
    }
    types = {
        name: {
            prop.get("name") for prop in item.get("properties", []) if prop.get("name")
        } | {"id", "title", "kind"}
        for name, item in type_items.items()
        if item.get("kind") != "union"
    }
    aliases = {
        item.get("alias"): item.get("type_name")
        for item in reference.get("profile_aliases", [])
        if item.get("alias") and item.get("type_name") in types
    }
    known = set(GENERIC_VARIABLES) | set(aliases)
    profile = reference.get("profile") or {}
    known.update(filter(None, (reference.get("root_alias"), profile.get("other_alias"))))
    known.update(
        item["name"] for item in BUILTIN_COMPATIBILITY_COLLECTIONS.get(dsl_key, [])
    )
    return {"types": types, "aliases": aliases, "known": known}


def _suggest(symbol, choices):
    match = difflib.get_close_matches(symbol, sorted(choices), n=1, cutoff=0.6)
    return match[0] if match else None


def _finding(severity, category, message, symbol, suggestion=None, location=None):
    return {
        "severity": severity,
        "category": category,
        "message": message,
        "symbol": symbol,
        "suggestion": suggestion,
        "location": location,
    }


def _check_top_level(name, inventory, findings, location):
    if name in inventory["known"]:
        return
    suggestion = _suggest(name, inventory["known"])
    findings.append(_finding(
        "warning", "dsl", "Unknown template variable `%s`." % name, name,
        suggestion, location,
    ))


def _by_type_name(node):
    from jinja2 import nodes

    if (
        isinstance(node, nodes.Getitem)
        and isinstance(node.node, nodes.Name)
        and node.node.name == "by_type"
        and isinstance(node.arg, nodes.Const)
        and isinstance(node.arg.value, str)
    ):
        return node.arg.value
    return None


def _location(source, node):
    if source.get("line_numbers") and node is not None and getattr(node, "lineno", None):
        return "line %s" % node.lineno
    return source.get("location")


def _check_filters_and_tests(ast, env, findings, source):
    from jinja2 import nodes

    for node in ast.find_all(nodes.Filter):
        if node.name not in env.filters:
            suggestion = _suggest(node.name, env.filters)
            findings.append(_finding(
                "warning", "filter", "Unknown Jinja filter `%s`." % node.name,
                node.name, suggestion, _location(source, node),
            ))
            # meta.find_undeclared_variables compiles the AST and otherwise
            # raises TemplateAssertionError before DSL checks can run.
            env.filters[node.name] = lambda value, *args, **kwargs: value
    for node in ast.find_all(nodes.Test):
        if node.name not in env.tests:
            suggestion = _suggest(node.name, env.tests)
            findings.append(_finding(
                "warning", "test", "Unknown Jinja test `%s`." % node.name,
                node.name, suggestion, _location(source, node),
            ))
            env.tests[node.name] = lambda value, *args, **kwargs: False


def _walk(node, bindings, inventory, findings, source):
    from jinja2 import nodes

    if isinstance(node, nodes.Getitem):
        type_name = _by_type_name(node)
        if type_name and type_name not in inventory["types"]:
            findings.append(_finding(
                "warning", "dsl", "Unknown `by_type` key `%s`." % type_name,
                type_name, _suggest(type_name, inventory["types"]),
                _location(source, node),
            ))

    if isinstance(node, nodes.For):
        child_bindings = dict(bindings)
        type_name = _by_type_name(node.iter)
        if isinstance(node.iter, nodes.Name):
            type_name = inventory["aliases"].get(node.iter.name)
        if (
            isinstance(node.target, nodes.Name)
            and type_name in inventory["types"]
            and not _assigns_name(node.body, node.target.name)
        ):
            child_bindings[node.target.name] = type_name
        for child in node.body:
            _walk(child, child_bindings, inventory, findings, source)
        for child in node.else_:
            _walk(child, bindings, inventory, findings, source)
        _walk(node.iter, bindings, inventory, findings, source)
        return

    if isinstance(node, nodes.Getattr) and isinstance(node.node, nodes.Name):
        type_name = bindings.get(node.node.name)
        if type_name and node.attr not in inventory["types"][type_name]:
            symbol = "%s.%s" % (node.node.name, node.attr)
            findings.append(_finding(
                "info", "dsl",
                "Possible field typo `%s` for type `%s`." % (node.attr, type_name),
                symbol, _suggest(node.attr, inventory["types"][type_name]),
                _location(source, node),
            ))

    for child in node.iter_child_nodes():
        _walk(child, bindings, inventory, findings, source)


def _assigns_name(nodes_to_check, name):
    """Treat reassigned loop variables as untyped for the whole loop body."""
    from jinja2 import nodes

    for candidate in nodes_to_check:
        if (
            isinstance(candidate, nodes.Assign)
            and isinstance(candidate.target, nodes.Name)
            and candidate.target.name == name
        ):
            return True
        if _assigns_name(candidate.iter_child_nodes(), name):
            return True
    return False

import base64
import io

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class BacklogImportWizard(models.TransientModel):
    _name = 'itlingo.backlog.import.wizard'
    _description = 'Import Backlog from Excel'

    project_id = fields.Many2one(
        'project.project', required=True, string='Project',
    )
    sprint_id = fields.Many2one(
        'itlingo.sprint', string='Target Sprint',
        domain="[('project_id', '=', project_id)]",
    )
    import_file = fields.Binary(string='Excel File', required=True)
    import_filename = fields.Char()
    result_message = fields.Text(readonly=True)

    def action_import(self):
        self.ensure_one()
        if load_workbook is None:
            raise UserError(_("openpyxl is required for Excel import."))
        if not self.import_file:
            raise UserError(_("Please select a file to import."))

        fp = io.BytesIO(base64.b64decode(self.import_file))
        wb = load_workbook(fp, read_only=True)
        ws = wb.active

        BacklogItem = self.env['itlingo.backlog.item']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0

        for row in rows:
            if not row or not row[1]:
                continue
            vals = {
                'name': str(row[1]),
                'item_type': str(row[2]).lower() if row[2] else 'story',
                'story_points': float(row[3]) if row[3] else 0,
                'priority': str(row[4]).lower() if row[4] else 'medium',
                'status': str(row[5]).lower() if row[5] else 'new',
                'description': str(row[7]) if len(row) > 7 and row[7] else '',
                'project_id': self.project_id.id,
            }
            if self.sprint_id:
                vals['sprint_id'] = self.sprint_id.id
            BacklogItem.create(vals)
            created += 1

        self.result_message = _('%d backlog items imported successfully.', created)

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
